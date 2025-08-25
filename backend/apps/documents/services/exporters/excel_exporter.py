import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from markdown_it import MarkdownIt

def export_document_to_excel(django_doc, include_comments=False):
    """
    Export a Django Document model instance to Excel (.XLSX) format with proper formatting.
    
    Returns:
        io.BytesIO: Excel file in .XLSX format ready for download/response
    """

    # Ensure we got the right type of object
    from apps.documents.models import Document  # your Django model
    if not isinstance(django_doc, Document):
        raise TypeError(f"Expected Django Document model, got {type(django_doc)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Document Content"
    
    # Set up Excel worksheet with better formatting
    ws.page_setup.orientation = 'portrait'
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    
    # Initialize markdown parser with table support
    md = MarkdownIt("commonmark").enable("table").disable("linkify")
    
    # Track current row
    current_row = 1
    
    # Add title with better formatting
    ws.cell(row=current_row, column=1, value="AI Concept Note")
    title_cell = ws[f"A{current_row}"]
    title_cell.font = Font(name='Arial', size=18, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Merge cells for title (spanning 5 columns for better appearance)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 2
    
    # Track all citations to avoid duplicates
    all_citations = set()
    
    # Process each section
    for sec in django_doc.sections.order_by("order"):
        content = sec.get_content() or ""
        
        # Process the content
        current_row = process_markdown_content(ws, content, md, current_row)
        
        # Collect unique citations
        for citation in sec.citations.all():
            all_citations.add(citation.reference_text)
        
        # Add comments if requested
        if include_comments:
            comments = sec.comments.filter(resolved=False)
            if comments:
                current_row += 1  # Add space
                ws.cell(row=current_row, column=1, value="Comments:")
                comment_header = ws[f"A{current_row}"]
                comment_header.font = Font(bold=True, size=12)
                current_row += 1
                
                for cm in comments:
                    ws.cell(row=current_row, column=1, value=f"• {cm.author}: {cm.content}")
                    comment_cell = ws[f"A{current_row}"]
                    comment_cell.font = Font(italic=True, size=10)
                    current_row += 1
                current_row += 1  # Add space after comments
    
    # Add references section if there are citations
    if all_citations:
        current_row += 2  # Add some space
        ws.cell(row=current_row, column=1, value="References")
        ref_header = ws[f"A{current_row}"]
        ref_header.font = Font(size=14, bold=True)
        current_row += 1
        
        # Sort citations alphabetically
        sorted_citations = sorted(all_citations)
        for citation_text in sorted_citations:
            cleaned_citation = clean_reference_text(citation_text)
            ws.cell(row=current_row, column=1, value=cleaned_citation)
            current_row += 1
    
    # Auto-adjust column widths
    adjust_column_widths(ws)
    
    # Save to BytesIO
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def process_markdown_content(ws, content, md, start_row):
    """Process markdown content and convert to Excel cells."""
    tokens = md.parse(content)
    current_row = start_row
    
    i = 0
    while i < len(tokens):
        token = tokens[i]
        
        if token.type == 'heading_open':
            level = int(token.tag[1])  # Extract number from h1, h2, etc.
            i += 1
            if i < len(tokens) and tokens[i].type == 'inline':
                ws.cell(row=current_row, column=1, value=tokens[i].content)
                heading_cell = ws[f"A{current_row}"]
                
                # Style based on heading level with professional fonts and colors
                if level == 1:
                    heading_cell.font = Font(name='Arial', size=16, bold=True, color="1F4E79")
                elif level == 2:
                    heading_cell.font = Font(name='Arial', size=14, bold=True, color="2F5597")
                else:
                    heading_cell.font = Font(name='Arial', size=12, bold=True, color="4472C4")
                
                current_row += 1
            i += 1  # Skip heading_close
            
        elif token.type == 'paragraph_open':
            i += 1
            if i < len(tokens) and tokens[i].type == 'inline':
                content_text = tokens[i].content
                if content_text.strip():  # Only add non-empty paragraphs
                    processed_text = process_inline_formatting(content_text)
                    cell = ws.cell(row=current_row, column=1, value=processed_text)
                    cell.font = Font(name='Arial', size=11)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    current_row += 1
            i += 1  # Skip paragraph_close
            
        elif token.type == 'table_open':
            current_row = process_table(ws, tokens, i, current_row)
            # Skip to table_close
            while i < len(tokens) and tokens[i].type != 'table_close':
                i += 1
            i += 1  # Skip table_close
            
        elif token.type == 'blockquote_open':
            i += 1
            quote_text = ""
            while i < len(tokens) and tokens[i].type != 'blockquote_close':
                if tokens[i].type == 'inline':
                    quote_text += tokens[i].content + " "
                i += 1
            
            if quote_text.strip():
                ws.cell(row=current_row, column=1, value=quote_text.strip())
                quote_cell = ws[f"A{current_row}"]
                quote_cell.font = Font(italic=True)
                quote_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                current_row += 1
            i += 1  # Skip blockquote_close
            
        elif token.type in ['bullet_list_open', 'ordered_list_open']:
            is_ordered = token.type == 'ordered_list_open'
            i += 1
            list_counter = 1
            
            while i < len(tokens) and tokens[i].type not in ['bullet_list_close', 'ordered_list_close']:
                if tokens[i].type == 'list_item_open':
                    i += 1
                    if i < len(tokens) and tokens[i].type == 'paragraph_open':
                        i += 1
                        if i < len(tokens) and tokens[i].type == 'inline':
                            list_text = tokens[i].content
                            if is_ordered:
                                prefix = f"{list_counter}. "
                                list_counter += 1
                            else:
                                prefix = "• "
                            
                            processed_text = process_inline_formatting(list_text)
                            ws.cell(row=current_row, column=1, value=prefix + processed_text)
                            current_row += 1
                i += 1
            i += 1  # Skip list_close
            
        else:
            i += 1
    
    return current_row + 1  # Add space after content

def process_table(ws, tokens, start_index, start_row):
    """Process a markdown table and convert to Excel table."""
    i = start_index + 1  # Skip table_open
    
    # Find table structure
    headers = []
    rows = []
    current_row_data = []
    
    while i < len(tokens) and tokens[i].type != 'table_close':
        token = tokens[i]
        
        if token.type == 'th_open':
            i += 1
            if i < len(tokens) and tokens[i].type == 'inline':
                headers.append(process_inline_formatting(tokens[i].content))
            i += 1  # Skip th_close
            
        elif token.type == 'td_open':
            i += 1
            if i < len(tokens) and tokens[i].type == 'inline':
                current_row_data.append(process_inline_formatting(tokens[i].content))
            i += 1  # Skip td_close
            
        elif token.type == 'tr_close':
            if current_row_data:
                rows.append(current_row_data)
                current_row_data = []
            i += 1
            
        else:
            i += 1
    
    # Create the table in Excel with improved styling
    if headers:
        current_row = start_row
        
        # Add headers with professional styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
        
        # Add data rows with alternating colors
        for row_idx, row_data in enumerate(rows):
            # Alternate row colors for better readability
            if row_idx % 2 == 0:
                fill_color = "F2F2F2"
            else:
                fill_color = "FFFFFF"
                
            for col, cell_data in enumerate(row_data[:len(headers)], 1):
                cell = ws.cell(row=current_row, column=col, value=cell_data)
                cell.font = Font(name='Arial', size=10)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            current_row += 1
        
        # Add borders to the table
        add_table_borders(ws, start_row, current_row - 1, len(headers))
        
        return current_row + 1  # Add space after table
    
    return start_row

def process_inline_formatting(text):
    """Process inline text formatting and return clean text (Excel doesn't support rich text easily)."""
    # Remove bold markers
    cleaned = re.sub(r'\*\*([^*]+?)\*\*', r'\1', text)
    # Remove italic markers  
    cleaned = re.sub(r'\*([^*]+?)\*', r'\1', cleaned)
    # Remove escaped asterisks
    cleaned = cleaned.replace('\\*', '*')
    return cleaned

def add_table_borders(ws, start_row, end_row, num_cols):
    """Add borders to a table range."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).border = thin_border

def adjust_column_widths(ws):
    """Auto-adjust column widths based on content with reasonable limits."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            if cell.value:
                # Convert to string and get length, accounting for line breaks
                cell_value = str(cell.value)
                # For cells with line breaks, take the longest line
                lines = cell_value.split('\n')
                cell_length = max(len(line) for line in lines) if lines else len(cell_value)
                
                if cell_length > max_length:
                    max_length = cell_length
        
        # Set width with padding, but use reasonable limits for Excel
        if max_length < 10:
            adjusted_width = 15
        elif max_length < 30:
            adjusted_width = max_length + 5
        elif max_length < 60:
            adjusted_width = max_length + 3
        else:
            adjusted_width = 80  # Max width for very long content
            
        ws.column_dimensions[column_letter].width = adjusted_width

def clean_reference_text(text):
    """Clean up reference text by removing duplicate chunks and formatting properly."""
    # Remove "Knowledge Base:" prefixes and chunk information
    cleaned = re.sub(r'Knowledge Base: [^(]+\(Chunk \d+\)\s*', '', text)
    
    # Remove extra whitespace
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    return cleaned if cleaned else text